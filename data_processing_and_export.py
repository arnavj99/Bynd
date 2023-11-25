import pandas as pd
import openpyxl
import numpy as np
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import PatternFill, Font

def get_available_metrics(dataframes):
    metrics = set()
    for df in dataframes:
        # This print statement is for debugging purposes
        print(f"Currently processing: {df}, Type: {type(df)}")
        # Ensure that df is a pandas DataFrame or Series
        if isinstance(df, (pd.DataFrame, pd.Series)):
            metrics.update(df.index.tolist())
        else:
            print(f"Error: Object {df} is not a pandas DataFrame or Series.")
    return list(metrics)

def create_metrics_mapping(dataframes):
    metrics_mapping = {}
    for df_name, df in dataframes.items():
        for metric in df.index:
            metrics_mapping[metric] = df_name
    return metrics_mapping

def fetch_selected_metrics(ticker, selected_metrics, dataframes, metrics_mapping, years):
    # Initialize DataFrame with ticker as index and metrics as columns
    data = pd.DataFrame(index=[ticker], columns=[f"{metric}_{'TTM' if year == 0 else 2023 - year}" for year in range(years) for metric in selected_metrics])
    for metric in selected_metrics:
        if metric in metrics_mapping:
            df_name = metrics_mapping[metric]
            df = dataframes[df_name]
            if metric in df.index:
                print(df.loc[metric])
                values = df.loc[metric].values
                values = np.concatenate([values[:1], values[2:]])  # Exclude the 2023 data
                for year in range(years):
                    try:
                        value = values[year]  # Get the value for the specified year
                    except IndexError:
                        value = pd.NA  # Fill with NA if the data for the year is not available

                    # Define column name based on the year
                    if year == 0:
                        column_name = f"{metric}_TTM"
                    else:
                        column_name = f"{metric}_{2023-year}"
                        
                    # Check if the value is a number and greater than 1000
                    if isinstance(value, (int, float)) and abs(value) > 100000:
                        value /= 1e6  # Convert to millions
                    elif isinstance(value, str):  # Check if the value is a string
                        try:
                            value = float(value)  # Try converting the string to a float
                            if abs(value) > 100000:
                                value /= 1e6  # Convert to millions
                        except ValueError:
                            pass  # Ignore if the string cannot be converted to a float

                    data.loc[ticker, column_name] = value  # Store the value in the DataFrame
            else:
                for year in range(years):
                    # Define column name based on the year
                    if year == 0:
                        column_name = f"{metric}_TTM"
                    else:
                        column_name = f"{metric}_{2023-year}"

                    data.loc[ticker, column_name] = pd.NA  # Fill with NA if the metric is not in the DataFrame
        else:
            for year in range(years):
                # Define column name based on the year
                if year == 0:
                    column_name = f"{metric}_TTM"
                else:
                    column_name = f"{metric}_{2023-year}"

                data.loc[ticker, column_name] = pd.NA  # Fill with NA if the metric is not in the metrics_mapping
    return data.reset_index().rename(columns={'index': 'ticker'})

def set_cell_format(worksheet, data, column):
    col_letter = openpyxl.utils.cell.get_column_letter(data.columns.get_loc(column) + 1)
    if pd.api.types.is_numeric_dtype(data[column]) and (data[column] <= 1).all():
        for row_num in range(2, len(data) + 2):
            cell = worksheet.cell(row=row_num, column=data.columns.get_loc(column) + 1)
            cell.number_format = '0.00%'
    return worksheet

def autosize_column(worksheet, data, column):
    col_letter = openpyxl.utils.cell.get_column_letter(data.columns.get_loc(column) + 1)
    max_length = max(len(item) for item in data[column].astype(str))
    worksheet.column_dimensions[col_letter].width = max_length + 2
    return worksheet

def format_header_row(worksheet):
    for cell in worksheet["1:1"]:
        cell.fill = openpyxl.styles.PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        if cell.value:
            cell.value = cell.value.title()
    return worksheet

def generate_excel(data, file_name='comparative_analysis.xlsx'):
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        data.to_excel(writer)
        worksheet = writer.sheets['Sheet1']
        for column in data.columns:
            if not pd.api.types.is_datetime64_any_dtype(data[column]):
                worksheet = set_cell_format(worksheet, data, column)
                worksheet = autosize_column(worksheet, data, column)
        worksheet = format_header_row(worksheet)
        writer.save()
    print(f'Data exported to {file_name}')
